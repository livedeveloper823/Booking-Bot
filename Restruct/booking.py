import openpyxl.workbook
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from datetime import date
from time import sleep
import func
import openpyxl
import json
import re
import os



def submit(accounts:list, urls:list):
    #  Today's month
    current_time = date.today()
    month = current_time.month

    password = "Book123456"
    for url in urls:
        for account in accounts:        
        
            remote_debugging_address = "127.0.0.1:9024"
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("debuggerAddress", remote_debugging_address)
            driver = webdriver.Chrome(options=chrome_options)

            driver.get("https://account.booking.com/register")
            func.wait_url(driver, "https://account.booking.com/register")
            try:
                driver.find_element(By.CSS_SELECTOR, "#onetrust-accept-btn-handler").click()
            except Exception as e:
                # Add your desired action if the element is not found
                print("Accept button not found. Performing alternative action...", e)
                pass
            sleep(2)

            if account["registered"] == False:
                sleep(2)
                func.find_element(driver, By.ID, "username").send_keys(account["email"])
                sleep(0.3)
                func.find_element(driver, By.CSS_SELECTOR, "#root > div > div > div.app > div.access-container.bui_font_body > div > div > div > div > div > div > div > form > div.Vw55_aACpYkyD1luIfis.cQj98l8L_Fc_IXYCWd4I.kBoHgdoepdx0bTMKriS3 > div:nth-child(2) > button").click()
                sleep(0.3)
                func.find_element(driver, By.ID, "new_password").send_keys(password)
                sleep(0.5)
                func.find_element(driver, By.ID, "confirmed_password").send_keys(password)
                sleep(0.5)
                func.find_element(driver, By.CSS_SELECTOR, "#root > div > div > div.app > div.access-container.bui_font_body > div > div > div > div > div > div > div > form > div > button").click()
                sleep(10)
                print("Successfully Registered! and Logined!")
                account["registered"] = True
                with open('accounts.json', 'w') as file:
                    json.dump(accounts, file)
                sleep(10)
                driver.get(url)
                sleep(10)
                try:
                    driver.find_element(By.CSS_SELECTOR, "#no_availability_msg > div.change_dates > div.c-next-available-dates--pp.c-next-available-dates--pp--www.c-next-available-dates--pp--www-distant > div.bui-carousel.bui-u-bleed\@small.c-next-available-dates__carousel-wrap > ul > li:nth-child(2) > a").click()
                    btn_element = func.find_element(driver, By.CSS_SELECTOR, "#bodyconstraint-inner > div:nth-child(8) > div > div.af5895d4b2 > div.df7e6ba27d > div.bcbf33c5c3 > div.dcf496a7b9.bb2746aad9 > div.d4924c9e74 > div.c82435a4b8.a178069f51.a6ae3c2b40.a18aeea94d.d794b7a0f7.f53e278e95.c6710787a4.bbefc5a07c > div.c066246e13 > div.c1edfbabcb > div > div.b1037148f8 > div.a4b53081e1 > div > div.da8b337763 > a")
                    link = btn_element.get_attribute("href")
                    driver.get(link)
                    driver.find_element(By.CLASS_NAME, "js-group-recommendation-reserve-btn").click()
                    sleep(1)
                    driver.find_element(By.CLASS_NAME, "js-reservation-button").click()
                    sleep(1)
                    func.find_element(driver, By.ID, "firstname").send_keys(account["first_name"])
                    sleep(0.5)
                    func.find_element(driver, By.ID, "lastname").send_keys(account["last_name"])
                    sleep(1)
                    func.find_element(driver, By.ID, "cc1").send_keys("Albania")
                    sleep(0.3)
                    func.find_element(driver, By.ID, "phone").send_keys("437481753")
                    sleep(1)
                    time = func.find_element(driver, By.CLASS_NAME, "bui-list__description").text
                    match_time = re.search(r'\d+:\d+', time)
                    if match_time:
                        # Extract the matched numbers
                        booking_time = match_time.group()
                        print(booking_time)  # Output: 14:00
                    else:
                        print("No numbers found")
                    sleep(0.5)
                    func.find_element(driver, By.ID, "checkin_eta_hour").send_keys(booking_time)
                    sleep(0.5)
                    func.find_element(driver, By.CLASS_NAME, "e2e-bp-submit-button--next-step").click()
                    sleep(2)
                    try:
                        func.find_element(driver, By.CSS_SELECTOR, "#cc_type").click()
                        sleep(1)
                        func.find_element(driver, By.CSS_SELECTOR, "#cc_number").send_keys("5425233430109903")
                        sleep(1)
                        card_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_type")
                        card = Select(card_element)
                        card.select_by_index(1) 
                        sleep(1)
                        month_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_month")
                        epire = Select(month_element)
                        epire.select_by_index(month-1)
                        func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                        sleep(5)
                    except:
                        func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                        sleep(5)
                    # # logout
                    # func.find_element(driver, By.CLASS_NAME, "bui-dropdown--end").click()
                    # sleep(0.3)
                    # logout = func.find_element(driver, By.CLASS_NAME, "bui-dropdown-menu__text")
                    # if logout.text == "Sign out":
                    #     logout.click()
                except:
                    try: 
                        driver.find_element(By.CLASS_NAME, "js-group-recommendation-reserve-btn").click()
                        sleep(1)
                        driver.find_element(By.CLASS_NAME, "js-reservation-button").click()
                        sleep(1)
                        func.find_element(driver, By.ID, "firstname").send_keys(account["first_name"])
                        sleep(0.5)
                        func.find_element(driver, By.ID, "lastname").send_keys(account["last_name"])
                        sleep(1)
                        func.find_element(driver, By.ID, "cc1").send_keys("Albania")
                        sleep(0.3)
                        func.find_element(driver, By.ID, "phone").send_keys("437481753")
                        sleep(1)
                        time = func.find_element(driver, By.CLASS_NAME, "bui-list__description").text
                        match_time = re.search(r'\d+:\d+', time)
                        if match_time:
                            # Extract the matched numbers
                            booking_time = match_time.group()
                            print(booking_time)  # Output: 14:00
                        else:
                            print("No numbers found")
                        sleep(0.5)
                        func.find_element(driver, By.ID, "checkin_eta_hour").send_keys(booking_time)
                        sleep(0.5)
                        func.find_element(driver, By.CLASS_NAME, "e2e-bp-submit-button--next-step").click()
                        sleep(2)
                        try:
                            func.find_element(driver, By.CSS_SELECTOR, "#cc_type").click()
                            sleep(1)
                            func.find_element(driver, By.CSS_SELECTOR, "#cc_number").send_keys("5425233430109903")
                            sleep(1)
                            card_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_type")
                            card = Select(card_element)
                            card.select_by_index(1) 
                            sleep(1)
                            month_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_month")
                            expire = Select(month_element)
                            expire.select_by_index(month-1)
                            func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                            sleep(5)
                        except:
                            func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                            sleep(5)
                    except:
                        pass
            if account["registered"] == True:
                sleep(2)
                func.find_element(driver, By.ID, "username").send_keys(account["email"])
                sleep(0.3)
                func.find_element(driver, By.CSS_SELECTOR, "#root > div > div > div.app > div.access-container.bui_font_body > div > div > div > div > div > div > div > form > div.Vw55_aACpYkyD1luIfis.cQj98l8L_Fc_IXYCWd4I.kBoHgdoepdx0bTMKriS3 > div:nth-child(2) > button").click()
                sleep(0.5)
                func.find_element(driver, By.CSS_SELECTOR, "#password").send_keys(password)
                sleep(0.3)
                func.find_element(driver, By.CSS_SELECTOR, "#root > div > div > div.app > div.access-container.bui_font_body > div > div > div > div > div > div > div > form > div > div:nth-child(3) > div > button").click()
                print("Successfully Logined!")
                sleep(10)
                driver.get(url)
                
                sleep(10)
                try:
                    driver.find_element(By.CSS_SELECTOR, "#no_availability_msg > div.change_dates > div.c-next-available-dates--pp.c-next-available-dates--pp--www.c-next-available-dates--pp--www-distant > div.bui-carousel.bui-u-bleed\@small.c-next-available-dates__carousel-wrap > ul > li:nth-child(2) > a").click()
                    btn_element = func.find_element(driver, By.CSS_SELECTOR, "#bodyconstraint-inner > div:nth-child(8) > div > div.af5895d4b2 > div.df7e6ba27d > div.bcbf33c5c3 > div.dcf496a7b9.bb2746aad9 > div.d4924c9e74 > div.c82435a4b8.a178069f51.a6ae3c2b40.a18aeea94d.d794b7a0f7.f53e278e95.c6710787a4.bbefc5a07c > div.c066246e13 > div.c1edfbabcb > div > div.b1037148f8 > div.a4b53081e1 > div > div.da8b337763 > a")
                    link = btn_element.get_attribute("href")
                    driver.get(link)
                    driver.find_element(By.CLASS_NAME, "js-group-recommendation-reserve-btn").click()
                    sleep(1)
                    driver.find_element(By.CLASS_NAME, "js-reservation-button").click()
                    sleep(1)
                    func.find_element(driver, By.ID, "firstname").send_keys(account["first_name"])
                    sleep(0.5)
                    func.find_element(driver, By.ID, "lastname").send_keys(account["last_name"])
                    sleep(1)
                    func.find_element(driver, By.ID, "cc1").send_keys("Albania")
                    sleep(0.3)
                    func.find_element(driver, By.ID, "phone").send_keys("437481753")
                    sleep(1)
                    time = func.find_element(driver, By.CLASS_NAME, "bui-list__description").text
                    match_time = re.search(r'\d+:\d+', time)
                    if match_time:
                        # Extract the matched numbers
                        booking_time = match_time.group()
                        print(booking_time)  # Output: 14:00
                    else:
                        print("No numbers found")
                    sleep(0.5)
                    func.find_element(driver, By.ID, "checkin_eta_hour").send_keys(booking_time)
                    sleep(0.5)
                    func.find_element(driver, By.CLASS_NAME, "e2e-bp-submit-button--next-step").click()
                    sleep(2)
                    
                    try:
                        func.find_element(driver, By.CSS_SELECTOR, "#cc_type").click()
                        sleep(1)
                        func.find_element(driver, By.CSS_SELECTOR, "#cc_number").send_keys("5425233430109903")
                        sleep(1)
                        card_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_type")
                        card = Select(card_element)
                        card.select_by_index(1) 
                        sleep(1)
                        month_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_month")
                        expire = Select(month_element)
                        expire.select_by_index(month-1)
                        func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                        sleep(5)
                    except:
                        func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                        sleep(5)
                    # # logout
                    # func.find_element(driver, By.CLASS_NAME, "bui-dropdown--end").click()
                    # sleep(0.3)
                    # logout = func.find_element(driver, By.CLASS_NAME, "bui-dropdown-menu__text")
                    # if logout.text == "Sign out":
                    #     logout.click()
                except:
                    try:
                        try:
                            driver.find_element(By.CLASS_NAME, "js-group-recommendation-reserve-btn").click()
                            sleep(1)
                            driver.find_element(By.CLASS_NAME, "js-reservation-button").click()
                            sleep(1)
                            func.find_element(driver, By.ID, "firstname").send_keys(account["first_name"])
                            sleep(0.5)
                            func.find_element(driver, By.ID, "lastname").send_keys(account["last_name"])
                            sleep(1)
                            func.find_element(driver, By.ID, "cc1").send_keys("Albania")
                            sleep(0.3)
                            func.find_element(driver, By.ID, "phone").send_keys("437481753")
                            sleep(1)
                            time = func.find_element(driver, By.CLASS_NAME, "bui-list__description").text
                            match_time = re.search(r'\d+:\d+', time)
                            if match_time:
                                # Extract the matched numbers
                                booking_time = match_time.group()
                                print(booking_time)  # Output: 14:00
                            else:
                                print("No numbers found")
                            sleep(0.5)
                            func.find_element(driver, By.ID, "checkin_eta_hour").send_keys(booking_time)
                            sleep(0.5)
                            func.find_element(driver, By.CLASS_NAME, "e2e-bp-submit-button--next-step").click()
                            sleep(2)
                            try:
                                func.find_element(driver, By.CSS_SELECTOR, "#cc_type").click()
                                sleep(1)
                                func.find_element(driver, By.CSS_SELECTOR, "#cc_number").send_keys("5425233430109903")
                                sleep(1)
                                card_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_type")
                                card = Select(card_element)
                                card.select_by_index(1) 
                                sleep(1)
                                month_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_month")
                                expire = Select(month_element)
                                expire.select_by_index(month-1)
                                func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                                sleep(5)
                            except:
                                func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                                sleep(5)
                        except:
                            apart_select = func.find_element(driver, By.CLASS_NAME, "hprt-nos-select")
                            apart = Select(apart_select)
                            apart.select_by_index(1)
                            driver.find_element(By.CLASS_NAME, "js-reservation-button").click()
                            sleep(1)
                            func.find_element(driver, By.ID, "firstname").send_keys(account["first_name"])
                            sleep(0.5)
                            func.find_element(driver, By.ID, "lastname").send_keys(account["last_name"])
                            sleep(1)
                            func.find_element(driver, By.ID, "cc1").send_keys("Albania")
                            sleep(0.3)
                            func.find_element(driver, By.ID, "phone").send_keys("437481753")
                            sleep(1)
                            time = func.find_element(driver, By.CLASS_NAME, "bui-list__description").text
                            match_time = re.search(r'\d+:\d+', time)
                            if match_time:
                                # Extract the matched numbers
                                booking_time = match_time.group()
                                print(booking_time)  # Output: 14:00
                            else:
                                print("No numbers found")
                            sleep(0.5)
                            func.find_element(driver, By.ID, "checkin_eta_hour").send_keys(booking_time)
                            sleep(0.5)
                            func.find_element(driver, By.CLASS_NAME, "e2e-bp-submit-button--next-step").click()
                            sleep(2)
                            try:
                                func.find_element(driver, By.CSS_SELECTOR, "#cc_type").click()
                                sleep(1)
                                func.find_element(driver, By.CSS_SELECTOR, "#cc_number").send_keys("5425233430109903")
                                sleep(1)
                                card_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_type")
                                card = Select(card_element)
                                card.select_by_index(1) 
                                sleep(1)
                                month_element = func.find_element(driver, By.CSS_SELECTOR, "#cc_month")
                                expire = Select(month_element)
                                expire.select_by_index(month-1)
                                func.find_element(driver, By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                                sleep(5)
                            except:
                                driver.find_element(By.CLASS_NAME,"js-bp-submit-button--complete-booking").click()
                                sleep(5)
                    except:
                        # func.find_element(driver, By.CSS_SELECTOR, "#basiclayout > div.hotelchars > div.page-section.hp--desc_highlights.js-k2-hp--block > div > div > div.bui-grid__column.bui-grid__column-4.k2-hp--highlights > div > div:nth-child(2) > div > div.ph-footer > button").click()
                        # sleep(1)
                        pass
                    # # logout
                    # func.find_element(driver, By.CLASS_NAME, "bui-dropdown--end").click()
                    # sleep(0.3)
                    # logout = func.find_element(driver, By.CLASS_NAME, "bui-dropdown-menu__text")
                    # if logout.text == "Sign out":
                    #     logout.click()          

if __name__ == "__main__":
    # import hotel urls
    excel = openpyxl.load_workbook("hotels.xlsx")
    sheet = excel.active
    rowCount = sheet.max_row + 1
    urls = []
    for row in range(1, rowCount):
        url = sheet.cell(row=row, column=1).value
        urls.append(url)
    print(urls)
    # import accounts
    accounts = []
    with open("accounts.json", "r") as file:
        accounts = json.load(file)
        print(accounts)
    
    submit(accounts, urls)
